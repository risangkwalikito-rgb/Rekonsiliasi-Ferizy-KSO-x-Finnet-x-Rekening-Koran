
from __future__ import annotations

from datetime import date
import io
import re
import unicodedata
from typing import Any

import pandas as pd
import streamlit as st


st.set_page_config(page_title="Rekonsiliasi Sharing Fee FINNET", layout="wide")


MASTER_FEE_DATA = [
    {"instrumen_pembayaran": "VA BRI", "status_pg": "Main", "admin_fee": 2220.0, "sharing_fee": 777.0, "pajak": 0.11, "sharing_fee_excl_tax": 700.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA BCA", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 278.0, "pajak": 0.11, "sharing_fee_excl_tax": 250.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA Mandiri", "status_pg": "Main", "admin_fee": 2220.0, "sharing_fee": 777.0, "pajak": 0.11, "sharing_fee_excl_tax": 700.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA BM", "status_pg": "Main", "admin_fee": 2220.0, "sharing_fee": 777.0, "pajak": 0.11, "sharing_fee_excl_tax": 700.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "DANA", "status_pg": "Secondary", "admin_fee": 0.015, "sharing_fee": 0.0012, "pajak": 0.11, "sharing_fee_excl_tax": 0.0011, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "OVO", "status_pg": "Secondary", "admin_fee": 0.0167, "sharing_fee": 0.0007, "pajak": 0.11, "sharing_fee_excl_tax": 0.0006, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA BSI", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 777.0, "pajak": 0.11, "sharing_fee_excl_tax": 700.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA Permata", "status_pg": "Main", "admin_fee": 2220.0, "sharing_fee": 666.0, "pajak": 0.11, "sharing_fee_excl_tax": 600.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "ShopeePay", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 310.0, "pajak": 0.11, "sharing_fee_excl_tax": 279.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "Link Aja", "status_pg": "Main", "admin_fee": 2220.0, "sharing_fee": 765.0, "pajak": 0.11, "sharing_fee_excl_tax": 689.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "Gopay", "status_pg": "Secondary", "admin_fee": 0.0167, "sharing_fee": 0.0, "pajak": 0.11, "sharing_fee_excl_tax": 0.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA CIMB", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 666.0, "pajak": 0.11, "sharing_fee_excl_tax": 600.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA BTN", "status_pg": "Main", "admin_fee": 2220.0, "sharing_fee": 888.0, "pajak": 0.11, "sharing_fee_excl_tax": 800.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA BJB", "status_pg": "Main", "admin_fee": 2220.0, "sharing_fee": 888.0, "pajak": 0.11, "sharing_fee_excl_tax": 800.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "Pospay", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 0.0, "pajak": 0.11, "sharing_fee_excl_tax": 0.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA Danamon", "status_pg": "Main", "admin_fee": 2220.0, "sharing_fee": 722.0, "pajak": 0.11, "sharing_fee_excl_tax": 650.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA Sumselbabel", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 722.0, "pajak": 0.11, "sharing_fee_excl_tax": 650.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA Maybank", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 722.0, "pajak": 0.11, "sharing_fee_excl_tax": 650.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "BLU", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 904.0, "pajak": 0.11, "sharing_fee_excl_tax": 814.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA Muamalat", "status_pg": "Main", "admin_fee": 2220.0, "sharing_fee": 722.0, "pajak": 0.11, "sharing_fee_excl_tax": 650.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "Indomaret", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 200.0, "pajak": 0.11, "sharing_fee_excl_tax": 180.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "Alfamart", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 200.0, "pajak": 0.11, "sharing_fee_excl_tax": 180.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA Maspion", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 0.0, "pajak": 0.11, "sharing_fee_excl_tax": 0.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA Nagari", "status_pg": "Main", "admin_fee": 2220.0, "sharing_fee": 722.0, "pajak": 0.11, "sharing_fee_excl_tax": 650.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA BTPN", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 0.0, "pajak": 0.11, "sharing_fee_excl_tax": 0.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA Neo Commerce", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 0.0, "pajak": 0.11, "sharing_fee_excl_tax": 0.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA Sinarmas", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 0.0, "pajak": 0.11, "sharing_fee_excl_tax": 0.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA Bank DKI", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 722.0, "pajak": 0.11, "sharing_fee_excl_tax": 650.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "VA BPD Jatim", "status_pg": "Main", "admin_fee": 2220.0, "sharing_fee": 722.0, "pajak": 0.11, "sharing_fee_excl_tax": 650.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "PT Pos Indonesia", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 0.0, "pajak": 0.11, "sharing_fee_excl_tax": 0.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "QRIS", "status_pg": "Secondary", "admin_fee": 0.007, "sharing_fee": 0.0013, "pajak": 0.11, "sharing_fee_excl_tax": 0.0011, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "Pegadaian", "status_pg": "Main", "admin_fee": 2220.0, "sharing_fee": 1500.0, "pajak": 0.11, "sharing_fee_excl_tax": 1351.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "Yomart", "status_pg": "Main", "admin_fee": 2220.0, "sharing_fee": 1500.0, "pajak": 0.11, "sharing_fee_excl_tax": 1351.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "Delima", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 200.0, "pajak": 0.11, "sharing_fee_excl_tax": 180.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "OctoClicks", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 0.0, "pajak": 0.11, "sharing_fee_excl_tax": 0.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "E-Pay BRI", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 0.0, "pajak": 0.11, "sharing_fee_excl_tax": 0.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "Permata", "status_pg": "Secondary", "admin_fee": 2220.0, "sharing_fee": 0.0, "pajak": 0.11, "sharing_fee_excl_tax": 0.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "Credit Card", "status_pg": "Secondary", "admin_fee": 0.015, "sharing_fee": 0.0, "pajak": 0.11, "sharing_fee_excl_tax": 0.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "Direct Debit", "status_pg": "Secondary", "admin_fee": 0.02, "sharing_fee": 0.0, "pajak": 0.11, "sharing_fee_excl_tax": 0.0, "settlement_hari_kerja": "H+1"},
    {"instrumen_pembayaran": "Finpay Code", "status_pg": "Main", "admin_fee": 2220.0, "sharing_fee": 1332.0, "pajak": 0.11, "sharing_fee_excl_tax": 1200.0, "settlement_hari_kerja": "H+1"},
]

ALIAS_MAP = {
    "va bri": "VA BRI",
    "bri va": "VA BRI",
    "briva": "VA BRI",
    "virtual account bri": "VA BRI",
    "va bca": "VA BCA",
    "bca va": "VA BCA",
    "virtual account bca": "VA BCA",
    "va mandiri": "VA Mandiri",
    "mandiri va": "VA Mandiri",
    "virtual account mandiri": "VA Mandiri",
    "mega va": "VA BM",
    "bank mega va": "VA BM",
    "dana": "DANA",
    "ovo": "OVO",
    "va bsi": "VA BSI",
    "vabsi": "VA BSI",
    "va permata": "VA Permata",
    "permata va": "VA Permata",
    "shopeepay": "ShopeePay",
    "shopee pay": "ShopeePay",
    "link aja": "Link Aja",
    "linkaja": "Link Aja",
    "tcash": "Link Aja",
    "gopay": "Gopay",
    "go pay": "Gopay",
    "va cimb": "VA CIMB",
    "vacimb": "VA CIMB",
    "va btn": "VA BTN",
    "vabtn": "VA BTN",
    "va bjb": "VA BJB",
    "vabjb": "VA BJB",
    "pospay": "Pospay",
    "va danamon": "VA Danamon",
    "vadanamon": "VA Danamon",
    "va sumselbabel": "VA Sumselbabel",
    "vasumselbabel": "VA Sumselbabel",
    "va maybank": "VA Maybank",
    "vamaybank": "VA Maybank",
    "blu": "BLU",
    "va muamalat": "VA Muamalat",
    "vamuamalat": "VA Muamalat",
    "indomaret": "Indomaret",
    "alfamart": "Alfamart",
    "va maspion": "VA Maspion",
    "vamaspion": "VA Maspion",
    "va nagari": "VA Nagari",
    "vanagari": "VA Nagari",
    "va btpn": "VA BTPN",
    "vabtpn": "VA BTPN",
    "va neo commerce": "VA Neo Commerce",
    "vaneocommerce": "VA Neo Commerce",
    "va sinarmas": "VA Sinarmas",
    "vasinarmas": "VA Sinarmas",
    "va bank dki": "VA Bank DKI",
    "vabankdki": "VA Bank DKI",
    "va bpd jatim": "VA BPD Jatim",
    "vabpdjatim": "VA BPD Jatim",
    "pt pos indonesia": "PT Pos Indonesia",
    "pt pos": "PT Pos Indonesia",
    "qris": "QRIS",
    "pegadaian": "Pegadaian",
    "yomart": "Yomart",
    "delima": "Delima",
    "octoclicks": "OctoClicks",
    "e-pay bri": "E-Pay BRI",
    "epay bri": "E-Pay BRI",
    "permata": "Permata",
    "credit card": "Credit Card",
    "direct debit": "Direct Debit",
    "finpay code": "Finpay Code",
    "finpay021": "Finpay Code",
    "va bni": "VA BNI",
    "vabni": "VA BNI",
    "bni va": "VA BNI",
}


def normalize_text(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[_\-/]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def format_number_id(value: Any, decimals: int = 2, trim_zero: bool = False) -> str:
    if value is None or pd.isna(value):
        return ""
    number = float(value)
    if trim_zero and number.is_integer():
        decimals = 0
    formatted = f"{number:,.{decimals}f}"
    formatted = formatted.replace(",", "_").replace(".", ",").replace("_", ".")
    if trim_zero and "," in formatted:
        formatted = formatted.rstrip("0").rstrip(",")
    return formatted


def format_integer_id(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return f"{int(round(float(value))):,}".replace(",", ".")


def parse_number(value: Any) -> float | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    text = text.replace("Rp", "").replace("rp", "").replace(" ", "")
    text = re.sub(r"[^0-9,.-]", "", text)

    if not text:
        return None

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif text.count(",") == 1 and "." not in text:
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")

    try:
        return float(text)
    except ValueError:
        return None


def compute_fee(amount: float | None, rule: float | None) -> float | None:
    if amount is None or pd.isna(amount) or rule is None or pd.isna(rule):
        return None
    if float(rule) < 1:
        return float(amount) * float(rule)
    return float(rule)


def normalize_date_range(
    value: Any,
    fallback_start: date,
    fallback_end: date,
) -> tuple[date, date]:
    if isinstance(value, (tuple, list)):
        if len(value) >= 2:
            start_date, end_date = value[0], value[1]
        elif len(value) == 1:
            start_date = end_date = value[0]
        else:
            start_date, end_date = fallback_start, fallback_end
    elif value:
        start_date = end_date = value
    else:
        start_date, end_date = fallback_start, fallback_end

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    return start_date, end_date


def clamp_date_range(
    start_date: date,
    end_date: date,
    min_date: date,
    max_date: date,
) -> tuple[date, date]:
    start_date = max(min_date, min(start_date, max_date))
    end_date = max(min_date, min(end_date, max_date))
    if start_date > end_date:
        return min_date, max_date
    return start_date, end_date


@st.cache_data(show_spinner=False)
def build_master_df() -> pd.DataFrame:
    master = pd.DataFrame(MASTER_FEE_DATA).copy()
    master["method_key"] = master["instrumen_pembayaran"].map(normalize_text)
    return master


def choose_excel_sheet(sheet_names: list[str]) -> str:
    for sheet_name in sheet_names:
        if normalize_text(sheet_name) == "detail settlement":
            return sheet_name
    return sheet_names[0]


@st.cache_data(show_spinner=False)
def read_uploaded_file(file_bytes: bytes, file_name: str) -> tuple[pd.DataFrame, str]:
    lower_name = file_name.lower()

    if lower_name.endswith(".csv"):
        for encoding in ("utf-8", "utf-8-sig", "latin1"):
            try:
                return pd.read_csv(io.BytesIO(file_bytes), encoding=encoding), "CSV"
            except Exception:
                continue
        raise ValueError("File CSV tidak bisa dibaca.")

    try:
        workbook = pd.ExcelFile(io.BytesIO(file_bytes))
    except Exception as exc:
        raise ValueError(f"File Excel tidak bisa dibaca: {exc}") from exc

    selected_sheet = choose_excel_sheet(workbook.sheet_names)

    try:
        dataframe = pd.read_excel(workbook, sheet_name=selected_sheet)
    except Exception as exc:
        raise ValueError(f"Sheet '{selected_sheet}' tidak bisa dibaca: {exc}") from exc

    return dataframe, selected_sheet


def detect_column_by_name(columns: list[str], target: str) -> str | None:
    target_key = normalize_text(target)
    for column in columns:
        if normalize_text(column) == target_key:
            return column
    return None


def get_merchant_amount_column(df: pd.DataFrame) -> tuple[str, str]:
    header_column = detect_column_by_name(df.columns.tolist(), "Merchant Amount")
    if header_column is not None:
        return header_column, "header"

    if df.shape[1] >= 21:
        return df.columns[20], "kolom U"

    raise ValueError("Kolom wajib tidak ditemukan: Merchant Amount / kolom U")


def parse_payment_datetime(series: pd.Series) -> pd.Series:
    if pd.api.types.is_datetime64_any_dtype(series):
        return pd.to_datetime(series, errors="coerce")

    if pd.api.types.is_numeric_dtype(series):
        parsed_serial = pd.to_datetime(series, unit="D", origin="1899-12-30", errors="coerce")
        if parsed_serial.notna().any():
            return parsed_serial

    text = series.astype(str).str.strip()

    for fmt in (
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y %I:%M %p",
    ):
        parsed = pd.to_datetime(text, format=fmt, errors="coerce")
        if parsed.notna().any():
            return parsed

    return pd.to_datetime(text, errors="coerce", dayfirst=False)


def resolve_payment_method(value: Any, master_df: pd.DataFrame) -> str | None:
    raw = normalize_text(value)
    if not raw:
        return None

    if raw in ALIAS_MAP:
        return ALIAS_MAP[raw]

    master_lookup = dict(zip(master_df["method_key"], master_df["instrumen_pembayaran"]))

    if raw in master_lookup:
        return master_lookup[raw]

    compact = raw.replace(" ", "")
    for key, canonical in master_lookup.items():
        if compact == key.replace(" ", ""):
            return canonical

    for key, canonical in master_lookup.items():
        if raw in key or key in raw:
            return canonical

    return None


def extract_branch_from_merchant_name(value: Any) -> str | None:
    if value is None or pd.isna(value):
        return None

    merchant_name = str(value).strip()
    if not merchant_name:
        return None

    merchant_name = re.sub(r"\s+", " ", merchant_name)
    return merchant_name


def prepare_dataset(df: pd.DataFrame) -> tuple[pd.DataFrame, str]:
    if df.shape[1] < 3:
        raise ValueError("Kolom C / Payment Date Time tidak ditemukan.")

    payment_method_column = detect_column_by_name(df.columns.tolist(), "Payment Method")
    if payment_method_column is None:
        raise ValueError("Kolom wajib tidak ditemukan: Payment Method")

    pg_provider_column = detect_column_by_name(df.columns.tolist(), "PG Provider")
    if pg_provider_column is None:
        raise ValueError("Kolom wajib tidak ditemukan: PG Provider")

    merchant_name_column = detect_column_by_name(df.columns.tolist(), "Merchant Name")
    if merchant_name_column is None:
        raise ValueError("Kolom wajib tidak ditemukan: Merchant Name")

    merchant_amount_column, merchant_amount_source = get_merchant_amount_column(df)

    working = df.copy()
    working["payment_datetime"] = parse_payment_datetime(working.iloc[:, 2])

    if working["payment_datetime"].notna().sum() == 0:
        raise ValueError(
            "Kolom C (Payment Date Time) tidak bisa diparse. Format yang didukung: m/dd/yyyy hh:mm:ss, mm/dd/yyyy hh:mm:ss, atau datetime native Excel."
        )

    working["payment_date"] = working["payment_datetime"].dt.date
    working["payment_method_raw"] = working[payment_method_column].astype(str).str.strip()
    working["pg_provider_raw"] = working[pg_provider_column].astype(str).str.strip()
    working["merchant_name_raw"] = working[merchant_name_column].astype(str).str.strip()
    working["merchant_amount"] = working[merchant_amount_column].map(parse_number)

    working = working.loc[working["payment_datetime"].notna()].copy()
    working = working.loc[working["payment_method_raw"].ne("")].copy()
    working = working.loc[working["merchant_name_raw"].ne("")].copy()
    working = working.loc[working["merchant_amount"].notna()].copy()
    working = working.loc[working["merchant_amount"] != 0].copy()
    working = working.loc[working["pg_provider_raw"].map(normalize_text).eq("finnet")].copy()

    if working.empty:
        raise ValueError("Tidak ada data FINNET yang valid setelah filter PG Provider, tanggal, Merchant Name, dan amount.")

    working["cabang"] = working["merchant_name_raw"].map(extract_branch_from_merchant_name)
    working = working.loc[working["cabang"].notna()].copy()

    if working.empty:
        raise ValueError("Kolom Merchant Name kosong setelah dibersihkan. Data cabang tidak bisa dibentuk.")

    return working, merchant_amount_source



def build_summary(prepared_df: pd.DataFrame, start_date: date, end_date: date) -> tuple[pd.DataFrame, pd.DataFrame]:
    master_df = build_master_df()

    filtered = prepared_df.loc[prepared_df["payment_date"].between(start_date, end_date)].copy()
    if filtered.empty:
        raise ValueError("Tidak ada data FINNET pada rentang tanggal yang dipilih.")

    filtered["payment_method"] = filtered["payment_method_raw"].map(lambda value: resolve_payment_method(value, master_df))

    unmatched_columns = ["cabang", "payment_method_raw"]
    rename_map = {"cabang": "Cabang", "payment_method_raw": "Payment Method Raw"}
    if "source_file" in filtered.columns:
        unmatched_columns.insert(0, "source_file")
        rename_map["source_file"] = "Source File"

    unmatched = (
        filtered.loc[filtered["payment_method"].isna(), unmatched_columns]
        .drop_duplicates()
        .rename(columns=rename_map)
        .sort_values(list(rename_map.values()), na_position="last")
        .reset_index(drop=True)
    )

    matched = filtered.loc[filtered["payment_method"].notna()].copy()
    if matched.empty:
        raise ValueError("Semua Payment Method pada data FINNET belum match ke master fee.")

    matched["trx_count"] = 1
    merged = matched.merge(
        master_df[["instrumen_pembayaran", "sharing_fee"]],
        how="left",
        left_on="payment_method",
        right_on="instrumen_pembayaran",
    )

    pivot = (
        merged.groupby(["cabang", "payment_method"], dropna=False, as_index=False)
        .agg(
            {
                "trx_count": "sum",
                "sharing_fee": "first",
            }
        )
        .sort_values(["cabang", "payment_method"], na_position="last")
        .reset_index(drop=True)
    )

    pivot["total_sharing_fee_excl_tax"] = pivot["trx_count"] * pivot["sharing_fee"]

    summary = pivot.rename(
        columns={
            "cabang": "Cabang",
            "payment_method": "Payment Method",
            "trx_count": "Jumlah Transaksi",
            "sharing_fee": "Master Sharing Fee",
            "total_sharing_fee_excl_tax": "Total Sharing Fee Exclude Tax",
        }
    )

    ordered_columns = [
        "Cabang",
        "Payment Method",
        "Jumlah Transaksi",
        "Master Sharing Fee",
        "Total Sharing Fee Exclude Tax",
    ]
    summary = summary[ordered_columns]

    return summary, unmatched


def load_uploaded_files(uploaded_files: list[Any]) -> tuple[pd.DataFrame | None, pd.DataFrame, list[str], date, date]:

    prepared_frames: list[pd.DataFrame] = []
    file_info_rows: list[dict[str, Any]] = []
    errors: list[str] = []

    for uploaded_file in uploaded_files:
        try:
            file_bytes = uploaded_file.getvalue()
            source_df, selected_sheet = read_uploaded_file(file_bytes, uploaded_file.name)
            prepared_df, merchant_amount_source = prepare_dataset(source_df)
            prepared_df = prepared_df.copy()
            prepared_df["source_file"] = uploaded_file.name
            prepared_frames.append(prepared_df)

            file_info_rows.append(
                {
                    "File": uploaded_file.name,
                    "Sheet": selected_sheet,
                    "Sumber Merchant Amount": merchant_amount_source,
                    "Min Tanggal": prepared_df["payment_date"].min(),
                    "Max Tanggal": prepared_df["payment_date"].max(),
                    "Jumlah Baris FINNET": len(prepared_df),
                }
            )
        except Exception as exc:
            errors.append(f"{uploaded_file.name}: {exc}")

    file_info_df = pd.DataFrame(file_info_rows)

    if not prepared_frames:
        return None, file_info_df, errors, date.today(), date.today()

    combined_df = pd.concat(prepared_frames, ignore_index=True)
    min_available_date = combined_df["payment_date"].min()
    max_available_date = combined_df["payment_date"].max()

    return combined_df, file_info_df, errors, min_available_date, max_available_date



def format_summary_for_display(summary_df: pd.DataFrame) -> pd.DataFrame:
    display_df = summary_df.copy()
    display_df["Jumlah Transaksi"] = display_df["Jumlah Transaksi"].map(format_integer_id)
    display_df["Master Sharing Fee"] = display_df["Master Sharing Fee"].map(
        lambda value: format_number_id(value, decimals=4, trim_zero=True)
    )
    display_df["Total Sharing Fee Exclude Tax"] = display_df["Total Sharing Fee Exclude Tax"].map(
        lambda value: format_number_id(value, decimals=2)
    )
    return display_df


def build_branch_sections(summary_df: pd.DataFrame) -> list[tuple[str, pd.DataFrame]]:
    sections: list[tuple[str, pd.DataFrame]] = []

    for cabang, group in summary_df.groupby("Cabang", dropna=False, sort=True):
        branch_table = group[
            [
                "Payment Method",
                "Jumlah Transaksi",
                "Master Sharing Fee",
                "Total Sharing Fee Exclude Tax",
            ]
        ].copy()

        subtotal = pd.DataFrame(
            [
                {
                    "Payment Method": "SUBTOTAL",
                    "Jumlah Transaksi": group["Jumlah Transaksi"].sum(),
                    "Master Sharing Fee": None,
                    "Total Sharing Fee Exclude Tax": group["Total Sharing Fee Exclude Tax"].sum(),
                }
            ]
        )
        branch_table = pd.concat([branch_table, subtotal], ignore_index=True)
        sections.append((str(cabang), branch_table))

    return sections


def render_branch_tables(summary_df: pd.DataFrame) -> None:
    for cabang, branch_table in build_branch_sections(summary_df):
        st.markdown(f"### {cabang}")
        st.dataframe(format_summary_for_display(branch_table), use_container_width=True, hide_index=True)

    grand_total = pd.DataFrame(
        [
            {
                "Payment Method": "GRAND TOTAL",
                "Jumlah Transaksi": summary_df["Jumlah Transaksi"].sum(),
                "Master Sharing Fee": None,
                "Total Sharing Fee Exclude Tax": summary_df["Total Sharing Fee Exclude Tax"].sum(),
            }
        ]
    )
    st.markdown("### GRAND TOTAL")
    st.dataframe(format_summary_for_display(grand_total), use_container_width=True, hide_index=True)


def to_excel_bytes(summary_df: pd.DataFrame, unmatched_df: pd.DataFrame, file_info_df: pd.DataFrame) -> bytes:
    from openpyxl.styles import Font

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        workbook = writer.book
        if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) == 1:
            workbook.remove(workbook["Sheet"])
        worksheet = workbook.create_sheet("Rekonsiliasi")
        writer.sheets["Rekonsiliasi"] = worksheet

        header_columns = [
            "Payment Method",
            "Jumlah Transaksi",
            "Master Sharing Fee",
            "Total Sharing Fee Exclude Tax",
        ]

        current_row = 1
        for cabang, branch_table in build_branch_sections(summary_df):
            worksheet.cell(row=current_row, column=1, value=cabang).font = Font(bold=True)
            current_row += 1

            for column_index, column_name in enumerate(header_columns, start=1):
                worksheet.cell(row=current_row, column=column_index, value=column_name).font = Font(bold=True)
            current_row += 1

            for _, row in branch_table.iterrows():
                for column_index, column_name in enumerate(header_columns, start=1):
                    worksheet.cell(row=current_row, column=column_index, value=row[column_name])
                current_row += 1

            current_row += 1

        worksheet.cell(row=current_row, column=1, value="GRAND TOTAL").font = Font(bold=True)
        current_row += 1
        for column_index, column_name in enumerate(header_columns, start=1):
            worksheet.cell(row=current_row, column=column_index, value=column_name).font = Font(bold=True)
        current_row += 1
        worksheet.cell(row=current_row, column=1, value="GRAND TOTAL")
        worksheet.cell(row=current_row, column=2, value=int(summary_df["Jumlah Transaksi"].sum()))
        worksheet.cell(row=current_row, column=4, value=float(summary_df["Total Sharing Fee Exclude Tax"].sum()))

        unmatched_df.to_excel(writer, index=False, sheet_name="Unmatched Method")
        file_info_df.to_excel(writer, index=False, sheet_name="Info File")

        branch_summary = (
            summary_df.groupby("Cabang", dropna=False, as_index=False)
            .agg(
                {
                    "Jumlah Transaksi": "sum",
                    "Total Sharing Fee Exclude Tax": "sum",
                }
            )
            .sort_values("Cabang", na_position="last")
            .reset_index(drop=True)
        )
        branch_summary.to_excel(writer, index=False, sheet_name="Ringkasan Cabang")

        for sheet_name, dataframe in {
            "Unmatched Method": unmatched_df,
            "Ringkasan Cabang": branch_summary,
            "Info File": file_info_df,
        }.items():
            sheet = writer.book[sheet_name]
            sheet.freeze_panes = "A2"
            for column_index, column_name in enumerate(dataframe.columns, start=1):
                values = [len(str(column_name))] + [len(str(value)) for value in dataframe[column_name].head(200).fillna("")]
                sheet.column_dimensions[sheet.cell(row=1, column=column_index).column_letter].width = min(max(values) + 2, 28)

        worksheet.freeze_panes = "A3"
        for column_index in range(1, 5):
            values = []
            for row in worksheet.iter_rows(min_col=column_index, max_col=column_index):
                for cell in row:
                    if cell.value is not None:
                        values.append(len(str(cell.value)))
            max_len = max(values) if values else 10
            worksheet.column_dimensions[worksheet.cell(row=1, column=column_index).column_letter].width = min(max_len + 2, 28)

    output.seek(0)
    return output.getvalue()


def render_metrics(summary_df: pd.DataFrame) -> None:
    total_cabang = int(summary_df["Cabang"].nunique()) if not summary_df.empty else 0
    total_method = int(summary_df["Payment Method"].nunique()) if not summary_df.empty else 0
    total_trx = int(summary_df["Jumlah Transaksi"].sum()) if not summary_df.empty else 0
    total_fee = float(summary_df["Total Sharing Fee Exclude Tax"].sum()) if not summary_df.empty else 0.0

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Jumlah Cabang", format_integer_id(total_cabang))
    col2.metric("Jumlah Payment Method", format_integer_id(total_method))
    col3.metric("Jumlah Transaksi", format_integer_id(total_trx))
    col4.metric("Total Sharing Fee Exclude Tax", format_number_id(total_fee, decimals=2))


def get_file_token(uploaded_files: Any) -> str:

    if not uploaded_files:
        return "__no_file__"

    if not isinstance(uploaded_files, list):
        uploaded_files = [uploaded_files]

    tokens = sorted(f"{uploaded_file.name}|{uploaded_file.size}" for uploaded_file in uploaded_files)
    return "||".join(tokens)


def main() -> None:
    st.title("Rekonsiliasi Sharing Fee FINNET")
    st.caption("Pilih rentang tanggal, upload satu atau beberapa file settlement FINNET, lalu proses rekonsiliasi pivot otomatis per cabang dengan header cabang terpisah berdasarkan Merchant Name.")

    today = date.today()
    st.session_state.setdefault("date_range_widget", (today, today))
    st.session_state.setdefault("loaded_file_token", "__no_file__")

    left_col, right_col = st.columns([1.2, 3.8], gap="large")

    with left_col:
        date_slot = st.container()
        uploader_slot = st.container()
        action_slot = st.container()
        info_slot = st.container()

    with uploader_slot:
        uploaded_files = st.file_uploader(
            "Settlement FINNET",
            type=["xlsx", "xls", "csv"],
            accept_multiple_files=True,
            help="Bisa upload beberapa file sekaligus. Jika ada sheet 'Detail Settlement', sheet itu yang dipakai. Jika tidak ada, app memakai sheet pertama.",
        )

    min_available_date = today
    max_available_date = today
    prepared_df: pd.DataFrame | None = None
    file_info_df = pd.DataFrame(columns=["File", "Sheet", "Sumber Merchant Amount", "Min Tanggal", "Max Tanggal", "Jumlah Baris FINNET"])
    load_errors: list[str] = []

    if uploaded_files:
        prepared_df, file_info_df, load_errors, min_available_date, max_available_date = load_uploaded_files(uploaded_files)

    current_file_token = get_file_token(uploaded_files)
    if current_file_token != st.session_state["loaded_file_token"]:
        st.session_state["loaded_file_token"] = current_file_token
        st.session_state["date_range_widget"] = (min_available_date, max_available_date)

    current_start, current_end = normalize_date_range(
        st.session_state["date_range_widget"],
        min_available_date,
        max_available_date,
    )
    current_start, current_end = clamp_date_range(
        current_start,
        current_end,
        min_available_date,
        max_available_date,
    )
    st.session_state["date_range_widget"] = (current_start, current_end)

    with date_slot:
        st.markdown("**Parameter Tanggal**")
        picked_range = st.date_input(
            "Rentang Payment Date Time",
            value=st.session_state["date_range_widget"],
            min_value=min_available_date,
            max_value=max_available_date,
            format="DD/MM/YYYY",
            key="date_range_widget",
        )

    start_date, end_date = normalize_date_range(picked_range, current_start, current_end)
    start_date, end_date = clamp_date_range(start_date, end_date, min_available_date, max_available_date)

    with action_slot:
        process_clicked = st.button("Proses Rekonsiliasi", type="primary", use_container_width=True)

    with info_slot:
        st.caption(f"Jumlah file terpilih: {len(uploaded_files) if uploaded_files else 0}")
        st.caption("Cabang dibaca dari kolom Merchant Name.")
        if not file_info_df.empty:
            unique_sources = sorted(file_info_df["Sumber Merchant Amount"].astype(str).unique().tolist())
            st.caption(f"Sumber Merchant Amount: {', '.join(unique_sources)}")

    with right_col:
        if not uploaded_files:
            st.info("Upload satu atau beberapa file settlement FINNET untuk mulai proses rekonsiliasi per cabang.")
            return

        if load_errors:
            for error_message in load_errors:
                st.error(error_message)

        if prepared_df is None or prepared_df.empty:
            st.warning("Belum ada file valid yang bisa diproses.")
            return

        st.markdown("**Info File Terbaca**")
        st.dataframe(file_info_df, use_container_width=True, hide_index=True)

        if not process_clicked:
            st.info("Pilih rentang tanggal terlebih dahulu, lalu klik **Proses Rekonsiliasi**.")
            return

        try:
            summary_df, unmatched_df = build_summary(prepared_df, start_date, end_date)
        except Exception as exc:
            st.error(str(exc))
            return

        st.markdown(
            f"**Periode Payment Date Time: {start_date.strftime('%d-%m-%Y')} s.d. {end_date.strftime('%d-%m-%Y')}**"
        )
        render_metrics(summary_df)
        render_branch_tables(summary_df)

        if not unmatched_df.empty:
            st.warning("Ada Payment Method yang belum match ke master fee pada sebagian cabang.")
            st.dataframe(unmatched_df, use_container_width=True, hide_index=True)

        st.download_button(
            "Download Hasil Rekonsiliasi Per Cabang",
            data=to_excel_bytes(summary_df, unmatched_df, file_info_df),
            file_name=f"rekonsiliasi_sharing_fee_finnet_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


if __name__ == "__main__":
    main()
